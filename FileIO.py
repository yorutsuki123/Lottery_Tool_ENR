import os
import openpyxl
import json

def loadXlsx(fn, sn, col):
    workbook = openpyxl.load_workbook(fn)
    sheet = workbook[sn]
    lst = []
    n = 2
    while sheet.cell(row=n, column=1).value != None:
        lst.append([])
        for i in range(col):
            lst[-1].append(sheet.cell(row=n, column=i+1).value)
        n = n + 1
    workbook.close()
    return lst

def dumpXlsx(fn, sn, ttl, width, data=[[]]):
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet(sn, 0)
    for i, t in enumerate(ttl):
        sheet.cell(row=1, column=1+i, value=t)
    for i, w in enumerate(width):
        sheet.column_dimensions[chr(65+i)].width = w
    for i, dd in enumerate(data):
        for j, d in enumerate(dd):
            sheet.cell(row=2+i, column=1+j, value=d)
    workbook.save(fn)
    workbook.close()

def loadJsonData(fn):
    d = []
    if not os.path.exists('./data'):
        os.makedirs('./data')
    if fn in os.listdir('./data/') and os.path.isfile('./data/' + fn):
        with open('./data/' + fn, "r") as f:
            d = json.load(f)
    else:
        with open('./data/' + fn, "w") as f:
            json.dump(d, f)
    return d

def dumpJsonData(fn, d):
    if not os.path.exists('./data'):
        os.makedirs('./data')
    with open('./data/' + fn, "w") as f:
        json.dump(d, f)